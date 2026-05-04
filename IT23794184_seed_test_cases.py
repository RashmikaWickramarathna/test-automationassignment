from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook

WORKBOOK_PATH = Path(__file__).resolve().parent / "Assignment 1 - Test cases.xlsx"
SHEET_NAME = "Test cases"

def length_type(text):
    if len(text) <= 30:
        return "S"
    elif len(text) <= 100:
        return "M"
    return "L"

CASES = [
("TC_001","oyata koheda inne dan?","Question forms","Interrogative sentence"),
("TC_002","oya monawada meyata kiyanne?","Question forms","Contains question format"),
("TC_003","ikmanata me weda iwara karanna","Command","Imperative instruction"),
("TC_004","mama gedhara yanavaa","Romanization / Spelling Variants","Vowel repetition"),
("TC_005","suba udasanak wewa oyata","Greeting","Common greeting phrase"),
("TC_006","mata meka therum ganna udaw karanna","Request","Polite request"),
("TC_007","oyata puluwannam meka hadala denna","Conditional Request","Uses 'puluwannam'"),
("TC_008","hari ehema karamu","Statement","Agreement expression"),
("TC_009","mata hithenne eya hari","Statement","Opinion expression"),
("TC_010","tika tika issarahata yanna","Command","Repeated phrasing"),
("TC_011","podi podi dewal walata bayawenna epa","Command","Negative instruction"),
("TC_012","ane! oyata meka therunada?","Question + Punctuation","Uses '?' and '!'"),
("TC_013","mama oyata kiyannawa","Statement","Simple statement"),
("TC_014","mn oyta kyannm","Romanization / Short form","Shortened words"),
("TC_015","api movie ekak balanna yamu","English Word Insertions","Word 'movie'"),
("TC_016","mama report eka submit karanna oni","English Word Insertions","report/submit"),
("TC_017","api already on the way inne","English Phrase","Full English phrase"),
("TC_018","oyage mobile data on da?","Question + English","Word 'data'"),
("TC_019","mama file eka upload kala","English Word Insertions","file/upload"),
("TC_020","mata telegram eken message ekak ewanna","English Word Insertions","telegram/message"),
("TC_021","api google meet eken kathakaramu","Proper Noun","Google Meet"),
("TC_022","pls meka ikmanata karanna","Abbreviation","'pls' used"),
("TC_023","id eka denna puluwanda","Question forms","Uses 'puluwanda'"),
("TC_024","ada test ekak thiyenawa","English Word Insertions","test"),
("TC_025","api next lec eka balamu","Abbreviation","next/lec"),
("TC_026","api galle walata yamu","Location","Place name"),
("TC_027","GALLE wala beach eka lassanai","Mixed Case + English","'beach' word"),
("TC_028","kasun ada school giya","Statement","Name + school"),
("TC_029","nishantha mata call kala","English Word Insertions","call"),
("TC_030","lamayi 10 denek awa","Numeric","Number usage"),
("TC_031","mama 3k aran awa","Numeric","Short numeric form"),
("TC_032","mata Rs 1500 denna","Currency","Rs usage"),
("TC_033","GBP 50 kiyanne kochcharada","Currency + Question","GBP usage"),
("TC_034","class eka 9am patan gannawa","Time + English","9am"),
("TC_035","api 6.30ta hamuwemu","Time Format","6.30"),
("TC_036","meeting eka 2026/05/10 thiyenawa","Date + English","Date format"),
("TC_037","mama 12 may ennam","Date","Month name"),
("TC_038","mama km 5k giya","Unit","km usage"),
("TC_039","uba patta wadak kala","Slang","'uba','patta'"),
("TC_040","oyata meka therunada bn?","Slang + Question","'bn' slang"),
]

def seed():
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET_NAME]

    for i, (tc_id, inp, typ, reason) in enumerate(CASES, start=2):
        ws.cell(i, 1).value = tc_id
        ws.cell(i, 2).value = length_type(inp)
        ws.cell(i, 3).value = inp
        ws.cell(i, 4).value = ""   # Expected output (you fill or automate)
        ws.cell(i, 5).value = ""   # Actual output
        ws.cell(i, 6).value = ""   # Status
        ws.cell(i, 7).value = typ
        ws.cell(i, 8).value = reason

    wb.save(WORKBOOK_PATH)
    print("✅ Your 40 test cases seeded successfully!")

if __name__ == "__main__":
    seed()